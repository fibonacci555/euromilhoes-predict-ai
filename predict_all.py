
import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from openpyxl import load_workbook
from keras.models import Sequential
from keras.layers import LSTM, Dense, Bidirectional, Dropout

def predict():

    wb = load_workbook("numeros.xlsx")
    ws = wb.active

    numeros = []

    for i in range(1337):
        edicao = [ws[f'B{str(2+i)}'].value,ws[f'C{str(2+i)}'].value,ws[f'D{str(2+i)}'].value,ws[f'E{str(2+i)}'].value,ws[f'F{str(2+i)}'].value,ws[f'G{str(2+i)}'].value,ws[f'H{str(2+i)}'].value ]
        numeros.append(edicao)

    df = pd.DataFrame(np.asmatrix(numeros))

    df.head()

    scaler = StandardScaler().fit(df.values)
    transformed_dataset = scaler.transform(df.values)
    transformed_df = pd.DataFrame(data=transformed_dataset,index=df.index)

    number_of_rows = df.values.shape[0]
    window_length = 7
    number_of_features = df.values.shape[1]

    train = np.empty([number_of_rows-window_length,window_length, number_of_features],dtype=float)
    label = np.empty([number_of_rows-window_length, number_of_features],dtype = float)
    window_length = 7

    for i in range(0,number_of_rows-window_length):
      train[i] = transformed_df.iloc[i:i+window_length,0:number_of_features]
      label[i] = transformed_df.iloc[i+window_length:i+window_length+1,0:number_of_features]


    model = Sequential()
    model.add(Bidirectional(LSTM(240,input_shape=(window_length,number_of_features),
                                 return_sequences=True)))
    model.add(Dropout(0.2))
    model.add(Bidirectional(LSTM(240,input_shape=(window_length,number_of_features),
                                 return_sequences=True)))
    model.add(Dropout(0.2))
    model.add(Bidirectional(LSTM(240,input_shape=(window_length,number_of_features),
                                 return_sequences=True)))
    model.add(Bidirectional(LSTM(240,input_shape=(window_length,number_of_features),
                                 return_sequences=False)))
    model.add(Dense(59))
    model.add(Dense(number_of_features))
    model.compile(loss="mse",optimizer="rmsprop",metrics=["accuracy"])

    model.fit(train,label,batch_size=100,epochs=300)

    ultimos = 7

    ultimos_euros = []

    for i in range(1337-ultimos,1337):
        edicao = [ws[f'B{str(2 + i)}'].value, ws[f'C{str(2 + i)}'].value, ws[f'D{str(2 + i)}'].value,
                  ws[f'E{str(2 + i)}'].value, ws[f'F{str(2 + i)}'].value, ws[f'G{str(2 + i)}'].value,
                  ws[f'H{str(2 + i)}'].value]
        ultimos_euros.append(edicao)



    to_predict =np.asarray(ultimos_euros)
    scaled_to_predict = scaler.transform(to_predict)


    scaled_to_predicted_output_1 = model.predict(np.array(np.array([scaled_to_predict])))
    prediction = scaler.inverse_transform(scaled_to_predicted_output_1).astype(int)[0]
    return prediction